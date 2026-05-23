"""Parse a .xlsx rate-card spreadsheet into a structured rate-card via LLM extraction.

Two-step: openpyxl reads the cells into a JSON-of-rows structure; AIService.complete_json
asks Claude to extract the rate card from that. The LLM call is the fuzzy part —
spreadsheets in the wild have arbitrary layouts, so structured extraction beats heuristics.
"""
from __future__ import annotations

import json
import logging
from io import BytesIO

from openpyxl import load_workbook

from app.services.llm import Tier, get_ai_service

logger = logging.getLogger(__name__)

MAX_BYTES = 5 * 1024 * 1024  # 5 MB

_EXTRACT_SYSTEM = (
    "You extract structured rate cards from spreadsheets for a creative agency."
    " Given the cells of one or more sheets, return STRICT JSON with this schema:"
    "\n"
    '  {"hourly_rates": {"role_key": rate_inr_int, ...},'
    ' "offerings": {"offering_key": {"name": "Display Name", "base_price": int}, ...},'
    ' "multipliers": {"key": percent_float, ...},'
    ' "low_confidence_fields": ["role.foo", "offering.bar"]}'
    "\n\n"
    "Keys are lowercase snake_case identifiers. Numeric values are INR rupees as ints"
    " unless the sheet clearly indicates a percentage. Leave any key you can't infer"
    " confidently out of the output; list it in low_confidence_fields with a brief locator."
    " Do not invent rates. Return empty objects rather than guessing."
)


def parse_workbook(content: bytes) -> list[dict]:
    """Read the workbook into a list of sheets, each with rows of cells."""
    wb = load_workbook(filename=BytesIO(content), data_only=True, read_only=True)
    sheets = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(values_only=True):
            cells = [{"value": v} for v in row]
            rows.append({"cells": cells})
        sheets.append({"name": sheet_name, "rows": rows})
    return sheets


async def extract_with_llm(sheets: list[dict]) -> dict:
    """Send the parsed sheets to Claude and ask for a structured rate card."""
    user_prompt = (
        "Spreadsheet contents:\n```json\n"
        + json.dumps(sheets, ensure_ascii=False, indent=2, default=str)[:50_000]
        + "\n```\n\nReturn the JSON described in the system prompt."
    )

    result = await get_ai_service().complete_json(
        prompt=user_prompt,
        system=_EXTRACT_SYSTEM,
        tier=Tier.BALANCED,
        max_tokens=2000,
    )

    if not isinstance(result, dict):
        raise ValueError("LLM returned a non-dict response")

    return {
        "hourly_rates": result.get("hourly_rates") or {},
        "offerings": result.get("offerings") or {},
        "multipliers": result.get("multipliers") or {},
        "low_confidence_fields": result.get("low_confidence_fields") or [],
    }


async def parse_and_extract(content: bytes) -> dict:
    """Top-level helper used by the upload endpoint."""
    sheets = parse_workbook(content)
    return await extract_with_llm(sheets)
